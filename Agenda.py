from openpyxl import Workbook

agenda = Workbook()##Creaza Documentul
excel_sheet = agenda.active##Activeaza/Se ia referinta sheet-ul din document
##Variabile
coloana = 1
rand = 1
##Variabile
def salveaza_agenda():
    agenda.save("agenda.xlsx")

def adauga_nume(nume):
    coloana = 1
    c_name = excel_sheet.cell(row=rand, column=coloana)
    c_name.value = str(nume)
    salveaza_agenda()

def adauga_numar(numar):
    global coloana
    coloana = 2
    global rand
    c_numar = excel_sheet.cell(row=rand, column=coloana)
    c_numar.value = str(numar)
    salveaza_agenda()
    rand = rand + 1
#def citeste_numele():

